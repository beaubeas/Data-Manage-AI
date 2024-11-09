from supercog.engine.tool_factory import ToolFactory, ToolCategory


from openpyxl import load_workbook
import pandas as pd
import re

from typing import Any, Callable

from supercog.engine.filesystem   import unrestricted_filesystem


class ExcelTool(ToolFactory):
    def __init__(self):
        super().__init__(
            id = "excel_connector",
            system_name = "Excel",
            logo_url=super().logo_from_domain("office.com"),
            auth_config = {},
            help="""
Read and write Excel files
""",
            category=ToolCategory.CATEGORY_FILES
        )

    def get_tools(self) -> list[Callable]:
        return self.wrap_tool_functions([
            self.read_excel_file,
            self.create_excel_from_csv,
            self.insert_rows,
            self.insert_cols,
            self.sheetnames,
            self.create_sheet,
            self.delete_rows,
            #self.access_excel_file,
        ])

    def read_excel_file(
        self,
        file_name,
        sheet_name='1'
        )  -> dict:
        """ Read all rows from an excel file
        file_name  -- full excel path
        """
        if re.match(r'\d+', sheet_name):
            sheet_name = int(sheet_name)-1

        df = pd.read_excel(file_name, sheet_name=sheet_name, engine='openpyxl')

        return self.get_dataframe_preview(df)

    def create_excel_from_csv(self, csv_filename, excel_filename, sheet_name='Sheet1') -> dict:
        """
        Creates a new Excel file from a CSV file.

        :param csv_filename: The name of the source CSV file
        :param excel_filename: The name of the Excel file to be created
        :param sheet_name: The name of the sheet in the new Excel file (default 'Sheet1')
        :return: A dictionary with status and message
        """
        try:
            with unrestricted_filesystem():
                df = pd.read_csv(csv_filename)
                with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                return {"status": "success", "message": f"Created {excel_filename} from {csv_filename}"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def insert_rows (
        self,
        file_name,
        start_index = 1,
        rows_cols_to_write = [],
        worksheet='Sheet1'
        )  -> dict:
        """ insert passed in rpws to sheet at start_index row
        file_name  -- full excel path
        """
        workbook = load_workbook(file_name)
        sheet = workbook[worksheet]
        rows_inserted = 0
        for index, row in enumerate(rows_cols_to_write, start=start_index):  # Start from desired insertion point
            sheet.insert_rows(index)
            rows_inserted += 1
            for col, value in enumerate(row, start=1):  # Column indexing starts at 1
                sheet.cell(row=index, column=col, value=value)
        with unrestricted_filesystem():
            workbook.save(file_name)
        print(f"Inserted {rows_inserted} rows to {file_name}.")
        return {"status": "success", "message": f"Inserted {rows_inserted} rows to {file_name}."}

    def insert_cols(
        self,
        file_name,
        start_index = 1,
        rows_cols_to_write = [],
        worksheet='Sheet1'
        )  -> dict:
        """ insert passed in columns to sheet at start_index col
        file_name  -- full excel path
        """
        workbook = load_workbook(file_name)
        sheet = workbook[worksheet]
        # Number of columns to insert
        num_columns = len(rows_cols_to_write)
    
        # Insert the new columns at the specified index
        sheet.insert_cols(start_index, amount=num_columns)
    
        # Populate the new columns with data
        for col_index, column in enumerate(rows_cols_to_write, start=start_index):
            for row_index, value in enumerate(column, start=1):
                cell = sheet.cell(row=row_index, column=col_index)
                cell.value = value
        with unrestricted_filesystem():
            workbook.save(file_name)
        print(f"Inserted {num_columns} cols to {file_name}.")
        return {"status": "success", "message": f"Inserted {num_columns} cols to {file_name}."}
    
    def sheetnames(
        self,
        file_name,
        )  -> dict:
        """ Return back a list of worksheets in the excel file
        file_name  -- full excel path
        """
        workbook = load_workbook(file_name)
        data = workbook.sheetnames
        return {"status": "success", "message": data}

    def create_sheet(
        self,
        file_name,
        worksheet='sheet1'
        )  -> dict:
        """ Create a new worksheet in the excel file
        file_name          -- full excel path
        worksheet='Sheet1' -- default to sheet1
        """
        workbook = load_workbook(file_name)
        with unrestricted_filesystem():
            workbook.create_sheet(title=worksheet)
            workbook.save(file_name)
        return {"status": "success", "message": f"added sheet {worksheet} to file {file_name}."}
    
    def delete_rows(
        self,
        file_name,
        start_index,
        worksheet='sheet1',
        amount=1
        )  -> dict:
        """  delete rows from an excel file start at start_index with amount to delete
        file_name  -- full excel path
        """
        workbook = load_workbook(file_name)
        sheet = workbook[worksheet]
        sheet.delete_rows(start_index, amount)
        return {"status": "success", "message": f"deleted {amount} rows from {worksheet} in file {file_name}."}
    
    def access_excel_file(
        self,
        file_name,
        command,
        message,
        start_index = 1,
        rows_cols_to_write = [],
        worksheet='sheet1',
        amount=1
        ) -> dict:
        """ full set of operations.
        Keyword arguments:
        file_name          -- full excel path
        command            -- read, read_row, insert_rows, sheetnames, create_sheet, delete_rows, insert_cols
        message            -- the user instructions that inspired this call
        start_index        -- for insert_rows or insert_cols the row  or col to start the insert on, 
                           -- for delete rows or cols the start row or col
        rows_cols_to_write -- a list of rows that will be inserted
        worksheet='Sheet1' -- default to sheet1
        amount=1           -- Number of rows or cols to delete

        Returns a preview of the dataframe which will hold the data.
        """
        workbook = load_workbook(file_name)
        
        #sheet = workbook.active  # This selects the active sheet. Alternatively: workbook['SheetName']
        if command  == 'read':
            return self.read_excel_file(file_name, sheet_name=worksheet)
        elif command  == 'insert_rows':
            return self.insert_rows(file_name,  start_index, rows_cols_to_write, worksheet)
        elif command  == 'insert_cols':
            return self.insert_cols(file_name,  start_index, rows_cols_to_write, worksheet)
        elif command  == 'sheetnames':
            data = workbook.sheetnames
            return {"status": "success", "message": data}
        elif command  == 'create_sheet':
            return self.create_sheet( file_name, worksheet)
        elif command  == 'delete_rows':
            self.delete_rows( file_name,  start_index, worksheet, amount)
            return {"status": "success", "message": f"{amount} rows deleted"}
        #else:
            #data = eval(command) # very powerful, but dangerous
        return {"status": "error", "message": f"Command {command} not recognized"}
