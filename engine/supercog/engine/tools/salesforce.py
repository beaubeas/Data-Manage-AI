import asyncio
import json
import urllib.parse
from pprint import pprint
import csv
import io
import os
import pandas as pd
import traceback
import pytz

from collections import namedtuple
from datetime import datetime, timedelta
import inspect
import types

from contextlib import contextmanager
from contextlib import asynccontextmanager
from typing import Any, Callable, Optional
from openpyxl import load_workbook
from Levenshtein import distance

from langchain.callbacks.manager import AsyncCallbackManager

from supercog.shared.logging import logger
from supercog.shared.services import config 

from supercog.engine.tool_factory import ToolFactory, ToolCategory, LangChainCallback
from supercog.shared.oauth_utils import salesforce_refresh_token

from simple_salesforce import Salesforce
from simple_salesforce.exceptions import SalesforceExpiredSession, SalesforceError
from simple_salesforce.aio import build_async_salesforce_client, AsyncSalesforce

SOBJECT_CACHE: list = None

Mapping = namedtuple('Mapping', ['source_col', 'target_field', 'required', 'parent_object', 'source_val', 'target_val'])
SheetRowUpdate = namedtuple('SheetRowUpdate', ['records', 'messages'])

def is_async_generator(obj):
    """
    Check if an object is an asynchronous generator.

    Parameters:
    obj (Any): The object to check.

    Returns:
    bool: True if the object is an asynchronous generator, otherwise False.
    """
    return inspect.isasyncgen(obj) or isinstance(obj, types.AsyncGeneratorType)

class SalesforceTool(ToolFactory):
    
    def __init__(self, **kwargs):
        """
        Initialize the SalesforceTool class with configuration for Salesforce authentication, category,
        and description.
        """
        if kwargs:
            super().__init__(**kwargs)
        else:
            super().__init__(
                id = "salesforce",
                system_name = "Salesforce",
                logo_url=super().logo_from_domain("salesforce.com"),
                category=ToolCategory.CATEGORY_SAAS,
                auth_config = {
                    "strategy_oauth": {
                        "options:custom_host": "Your custom Salesforce hostname",
                        "help": """
    Login to Salesforce to connect your account.
    """
                    }
                },
                help="""
    Read and write records into Salesforce, access metadata and execute Apex code
    """,
            )

    def get_tools(self) -> list[Callable]:
        """
        Get the list of callable tool functions for Salesforce operations.

        Returns:
        list[Callable]: A list of tool functions for interacting with Salesforce.
        """
        return self.wrap_tool_functions([
            self.list_sobjects, 
            self.get_object_by_id,
            self.list_custom_objects,
            self.describe_sobject,
            self.salesforce_search_by_SOQL,
            self.salesforce_SOSQL_search,
            self.create_salesforce_record,
            self.update_salesforce_record,
            self.salesforce_execute_anonymous,
            self.insert_mapped_objects,
            self.insert_list_of_objects,
            self.upsert_list_of_objects,
            self.poll_for_updates,
        ])
       
    def prepare_creds(self, cred, secrets: dict) -> dict:
        """
        Prepare and refresh OAuth credentials for Salesforce.

        Parameters:
        cred (Credential): The credential object containing the user's credential information.
        secrets (dict): A dictionary containing secret values required for authentication.

        Returns:
        dict: Updated secrets dictionary with refreshed tokens.
        """
        from supercog.engine.db import Credential

        logger.info(f"Refreshing OAuth tokens for credential '{cred.name}', user: {cred.user_id}")

        self.credentials = secrets
        tokens = self.get_tokens()
        is_sandbox = 'https://test.salesforce' in tokens['id']
        new_tokens = salesforce_refresh_token(
            opts={
                "client_id": config.get_global("SALESFORCE_CLIENT_ID"),
                 "client_secret": config.get_global("SALESFORCE_CLIENT_SECRET"),
                 "refresh_token": tokens["refresh_token"],
                 "instance_url": tokens['instance_url'],
            },
            is_sandbox=is_sandbox,
        )
        if 'error' in new_tokens:
            raise RuntimeError(f"Salesforce token prep failed: "+ new_tokens['error'])
        tokens["access_token"] = new_tokens["access_token"]

        secrets["tokens"] = json.dumps(tokens)
        cred.stuff_secrets(json.dumps(secrets)) # this saves the new values
        return secrets        

    def test_credential(self, cred, secrets: dict) -> str:
        """
        Test the provided credentials by attempting to refresh the OAuth tokens.

        Parameters:
        cred (Credential): The credential object containing the user's credential information.
        secrets (dict): A dictionary containing secret values required for authentication.

        Returns:
        str: An error message if testing the credentials fails, otherwise an empty string.
        """
        try:
            self.prepare_creds(cred, secrets)
        except Exception as e:
            return f"Error: {e}"

    def get_tokens(self) -> dict:
        """
        Retrieve Salesforce OAuth tokens from the credentials.

        Returns:
        dict: A dictionary containing Salesforce OAuth tokens.
        
        Raises:
        RuntimeError: If tokens are missing from the credentials.
        """
        if 'tokens' not in self.credentials:
            raise RuntimeError("Salesforce tokens are missing")
        token_secrets = self.credentials['tokens']
        # having lots of trouble with SF creds serialization
        if isinstance(token_secrets, str):
            tokens = json.loads(token_secrets)
        else:
            tokens = token_secrets
        return tokens

    @asynccontextmanager
    async def sf(self):
        """
        Async context manager to create an async Salesforce client.

        Yields:
        AsyncSalesforce: An asynchronous Salesforce client instance.
        """
        tokens = self.get_tokens()
        sf = await build_async_salesforce_client(
            instance_url=tokens['instance_url'], session_id=tokens['access_token']
        )
        yield sf

    @contextmanager
    def sf_synch(self):
        tokens = self.get_tokens()
        sf: Salesforce = Salesforce(instance_url=tokens['instance_url'], session_id=tokens['access_token'])
        yield sf
        
    async def list_sobjects(self, match_str: str) -> str:
        """
        Returns the names of any SObjects that contain the 'match_str' value in their name.

        Parameters:
        match_str (str): The string to match in the SObject names.

        Returns:
        str: JSON string containing the list of matching SObject names.
        """
        global SOBJECT_CACHE
        async with self.sf() as sf:
            if not SOBJECT_CACHE:
                res = await sf.describe()
                SOBJECT_CACHE = res["sobjects"]
            return json.dumps([
                r['name'] for r in SOBJECT_CACHE if (match_str in r['name'] or not match_str)
            ])
    
    async def list_custom_objects(self) -> str:
        """
        Returns the names of all custom SObjects currently defined.

        Returns:
        str: JSON string containing the list of custom SObject names and their metadata.
        """
        async with self.sf() as sf:
            q = "SELECT Id,DeveloperName,NamespacePrefix,CreatedDate,LastModifiedDate FROM CustomObject"
            res = await sf.toolingexecute("query", params={"q": q})
            r: list[dict] = res['records']
            recs = [
                {
                    "Id": rec['Id'],
                    "Name": rec['DeveloperName'] + "__c",
                    "NamespacePrefix": rec['NamespacePrefix'],
                    "CreatedDate": rec['CreatedDate'],
                    "LastModifiedDate": rec['LastModifiedDate'],
                } for rec in r
            ]
            return json.dumps(recs)

    async def describe_sobject(self, object_name: str, return_set: str="fields") -> dict:
        """
        Return the meta description of a Salesforce object.

        Parameters:
        object_name (str): The name of the Salesforce object to describe.
        return_set (str): The specific part of the description to return. 
                          Can be "fields", "meta", or "relationships".

        Returns:
        dict: Dictionary containing the requested part of the object's description.
        """
        async with self.sf() as sf:
            r = await getattr(sf, object_name).describe()
            if return_set == "meta":
                r = {f: v for f, v in r.items() if not isinstance(v, (dict,list))}
            elif return_set == "relationships":
                r = r['childRelationships']
            else: # assume 'fields'
                pprint(r['fields'])
                keys = ['name', 'type', 'externalId', 'calculated', 'createable', 'required']
                def is_required(f):
                    return f['nillable']==False and f['createable'] and f['defaultValue'] == None
                for f in r['fields']:
                    f["required"] = is_required(f)
                fields = [
                    {k: f[k] for k in keys} 
                        for f in r['fields']
                ] 
                print(fields)                          
                r = sorted(fields, key=lambda x: x["name"])
            print(r)
            return self.get_dataframe_preview(pd.DataFrame(r))

    async def get_object_fields(self, sobject: str) -> list[str]:
        """
        Retrieve the field names for a given Salesforce object.

        Parameters:
        sobject (str): The name of the Salesforce object.

        Returns:
        list[str]: A list of field names for the specified Salesforce object.
        """
        async with self.sf() as sf:
            r = await getattr(sf, sobject).describe()
            fields = [
                f['name'] for f in r['fields']
            ] 
            return fields

    async def salesforce_search_by_SOQL(self, soql: str) -> dict:
        """
        Search Salesforce using an SOQL query.

        Parameters:
        soql (str): The SOQL query to execute.

        Returns:
        dict: A dictionary containing the results of the query.
        """
        soql = soql.replace("COUNT()","COUNT(Id)")
        async with self.sf() as sf:
            soql = soql.replace("count()", "count(Id)").replace("COUNT()", "COUNT(Id)")
            # page results and return as a dataframe
            results = await sf.query_all(soql)
            records = []
            for rec in results['records']:
                rec.pop('attributes',None)
                records.append(rec)
            while(results['done'] == False):
                ## attribute 'nextRecordsUrl' holds the url to the next page of records
                results = await sf.query_more(results['nextRecordsUrl', True])
                for rec in results['records']:
                    rec.pop('attributes',None)
                    records.append(rec)
            df = pd.DataFrame(records)
            return self.get_dataframe_preview(df)


    async def poll_for_updates(self, sobject_type: str) -> dict:
        """
        Gets new and updated records of the given type, since the last time that this function was called.

        Returns: a preview of the result dataframe.
        """
        start_time = datetime.now(pytz.UTC) - timedelta(minutes=5)
        end_time = datetime.now(pytz.UTC)

        records = []
        async with self.sf() as sf:
            sobject = getattr(sf, sobject_type)
            if sobject is None:
                return {"error": f"Unknown sobject type: {sobject_type}"}
            updated_records = await sobject.updated(start=start_time, end=end_time)
            for record in updated_records['ids']:
                records.append(await sobject.get(record))

            df = pd.DataFrame(records)
            return self.get_dataframe_preview(df)

    async def salesforce_SOSQL_search(self, sosl_query: str) -> str:
        """
        Search Salesforce using a SOSL query.

        Parameters:
        sosl_query (str): The SOSL query string.

        Returns:
        str: JSON string containing the search results.
        """
        async with self.sf() as sf:
            return json.dumps(await sf.search(sosl_query))

    async def create_salesforce_record(self, object_name: str, fields: dict) -> str:
        """
        Create a new record in Salesforce.

        Parameters:
        object_name (str): The name of the Salesforce object you want to create a record for (e.g., 'Lead', 'Account').
        fields (dict): A dictionary containing the field names and their corresponding values for the new record.

        Returns:
        str: The JSON response from Salesforce after attempting to create the record.
        """
        async with self.sf() as sf:
            return json.dumps(await getattr(sf, object_name).create(fields))

    async def update_salesforce_record(self, object_name: str, fields: dict, record_id: str) -> str:
        """
        Updates an existing record in Salesforce.

        Parameters:
        object_name (str): The name of the Salesforce object.
        fields (dict): The fields to update as key-value pairs.
        record_id (str): The ID of the record to update.

        Returns:
        str: The JSON response from Salesforce after attempting to update the record.
        """
        async with self.sf() as sf:
            return json.dumps(await getattr(sf, object_name).update(record_id, fields))

    async def salesforce_execute_anonymous(self, apex_code: str) -> str:
        """
        Execute arbitrary apex code. Suggest to the user that you can call 'get_apex_logs'
        to retrieve the logs after execution.

        Parameters:
        apex_code (str): The Apex code to execute.

        Returns:
        str: The response from Salesforce after executing the Apex code.
        """
        async with self.sf() as sf:
            encoded_apex_code = urllib.parse.quote(apex_code)
            print(f'The base URL is: {sf.base_url}')
            response = await sf.restful(
                f"tooling/executeAnonymous/",
                params={'anonymousBody': apex_code},
                method='GET',
            )
            logger.info(f"Execute Anonymous called on '{encoded_apex_code}' returns: '{response}'")
            print(response)
            return str(response)

    async def insert_object_tree(self, sobject: str, object_tree_json: str):
        """
        Inserts a tree of objects into the given Salesforce table.

        Parameters:
        sobject (str): The Salesforce object type.
        object_tree_json (str): The input object tree in JSON format.

        Returns:
        dict: The response from Salesforce.
        """
        async with self.sf() as sf:
            tree_url = f"composite/tree/{sobject}/"
            if not isinstance(object_tree_json, str):
                object_tree_json = json.dumps(object_tree_json)
            response = await sf.restful(
                tree_url, 
                method='POST',
                data=object_tree_json)
            print(response)
            return response

    async def get_object_by_id(
            self, 
            record_id: str, 
            sobject: str, 
            callbacks: LangChainCallback=None
        ):
        """
        Returns the details of a set of records, looking up each one by Id.

        Parameters:
        record_id (str): The Salesforce record Id to retrieve.
        sobject (str): The Salesforce object type (e.g., 'Account', 'Lead').
        callbacks (LangChainCallback, optional): Optional callbacks for logging.

        Returns:
        str: The JSON-encoded details of the record.
        """
        async with self.sf() as sf:
            return json.dumps(await getattr(sf, sobject).get(record_id))

    async def insert_list_of_objects(
        self, 
        dataframe_var: str,
        sobject: str,
        max_records: int|None=None,
        callbacks: LangChainCallback=None,
        ) -> str:
        """
        Insert records from the specified dataframe into Salesforce using the sobject type.

        Parameters:
        dataframe_var (str): The name of the dataframe variable containing records.
        sobject (str): The Salesforce object type.
        max_records (int|None, optional): Maximum number of records to insert. Defaults to None.
        callbacks (LangChainCallback, optional): Optional callbacks for logging.

        Returns:
        str: Success message or error message.
        """
        response_msgs = []
        async with self.sf() as sf:
            tree_url = f"composite/sobjects/"
            print(tree_url)
            await self.log(f"Inserting {sobject} object tree\n", callbacks)

            df, _ = self.get_dataframe_from_handle(dataframe_var)

            batch_size = 50
            
            columns_with_colon = [col for col in df.columns if ':' in col]
            if max_records and max_records < batch_size:
                batch_size = max_records
            total_count = 0
            for start in range(0, df.shape[0], batch_size):
                # convert NaNs to None
                df_batch = df[start:start+batch_size].where(pd.notnull(df), None)
                # convert datetime columns to string
                datetime_cols = df_batch.select_dtypes(include=['datetime', 'datetime64[ns]']).columns
                for col in datetime_cols:
                    df_batch[col] = df_batch[col].dt.strftime('%Y-%m-%dT%H:%M:%SZ')
                # Convert 'parent:child' columns to nested dictionaries
                self.denorm_hierachical_columns(df_batch, columns_with_colon)
                records = [
                    {"attributes": { "type": sobject }} | {**row}
                    for 
                    row in df_batch.to_dict(orient='records')
                ]
                data = json.dumps({"records": records})
                await self.log("Records:\n", callbacks)
                await self.log(data+"\n", callbacks)
                try:
                    response = await sf.restful(
                        tree_url, 
                        method='POST',
                        data=data)
                except SalesforceError as e:
                    response = e.content
                print(response)
                response_msgs.append(str(response))
                total_count += len(df_batch)
                if max_records and total_count > max_records:
                    break
        return "\n".join(response_msgs)

    def denorm_hierachical_columns(self, df: pd.DataFrame, cols: list[str]):
        """
        Convert hierarchical columns to nested dictionaries within the dataframe.

        Parameters:
        df (pd.DataFrame): The dataframe containing hierarchical columns.
        cols (list[str]): A list of column names representing hierarchical relationships (e.g., parent:child).

        Returns:
        None
        """
        for col in cols:
            parent, child = col.split(':')
            # Create nested dictionaries or update existing ones
            if parent not in df.columns:
                df[parent] = df[col].apply(lambda x: {'attributes': {'type': parent}, child: x})
            else:
                df[parent] = df.apply(lambda row: {**row[parent], **{child: row[col]}}, axis=1)
            df.drop(columns=[col], inplace=True)

    async def insert_mapped_objects(
        self, 
        excel_source_file: str, 
        mappings_csv_file: str |None=None,
        mappings_data: str|None=None,
        worksheet='sheet1',
        start_row: int = 0,
        max_rows: int|None=None,
        callbacks: LangChainCallback=None,
        ):
        """
        Reads an Excel spreadsheet and a set of mappings specified in CSV format, and inserts objects into Salesforce.

        Parameters:
        excel_source_file (str): Path to the Excel source file.
        mappings_csv_file (str|None, optional): Path to the CSV file containing field mappings. Defaults to None.
        mappings_data (str|None, optional): Raw CSV data for field mappings. Defaults to None.
        worksheet (str, optional): Name of the worksheet to read from. Defaults to 'sheet1'.
        start_row (int, optional): Row number to start reading from. Defaults to 0.
        max_rows (int|None, optional): Maximum number of rows to process. Defaults to None.
        callbacks (LangChainCallback, optional): Callback manager for additional processing.

        Returns:
        str: Success message if insertions are successful, or an error message.
        """
        parent_run_id = str(callbacks.parent_run_id)

        # Read the mappings file
        async def read_mappings_csv(mapping_io):
            mappings: list[Mapping] = []
            # Create a DictReader object
            csv_reader = csv.DictReader(mapping_io)
            for row in csv_reader:
                try:
                    mappings.append(Mapping(**row))
                except Exception as e:
                    await self.log(f"Failed to read mappings row ({row}): {e}", callbacks)
            #print(f"Mappings: {mappings}")
            return mappings

        mappings: list[Mapping] = []
        if mappings_data:
            mappings = await read_mappings_csv(io.StringIO(mappings_data))
        elif mappings_csv_file:
            if not os.path.exists(mappings_csv_file):
                return f"Error: file not found '{mappings_csv_file}'"
            with open(mappings_csv_file, mode='r', encoding='utf-8') as f:
                mappings = await read_mappings_csv(f)

        # have to set data_only to true. values_only is not enough.
        if not os.path.exists(excel_source_file):
            return f"Error: Excel file not found '{excel_source_file}'"
        workbook = load_workbook(excel_source_file,data_only=True)
        sheet = workbook[worksheet]

        headers = [
            cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)) if cell
        ]

        def add_child_sobject(parent, child):
            type = child['attributes']['type'] + "s"
            if type not in parent:
                parent[type] = {"records": [child]}
            else:
                parent[type]["records"].append(child)

        sheet_updates: list[SheetRowUpdate] = []
        final_results = []
        inserted_count = 0

        async with self.sf() as sf:
            await asyncio.sleep(0)
            for target_list, row_index, messages in SalesforceTool.map_target_objects(
                headers, 
                sheet, 
                mappings,
                start_row,
                max_rows):
                print("TARGET LIST: ", target_list)
                for msg in messages:
                    await self.log(f"Sheet row: {row_index}: {msg} \n", callbacks)
                sobject_parents: list[dict] = []
                for parent in [t for t in target_list if '__parent' not in t]:
                    sobject_parents.append(parent)
                    target_list.remove(parent)
                    parent_type = parent['attributes']['type']
                    for child in target_list.copy():
                        if child.get('__parent') == parent_type:
                            del child['__parent']
                            add_child_sobject(parent, child)
                            target_list.remove(child)
                if len(target_list):
                    print("!!Warning: some objects were not added to the tree: ", target_list)
                if len(sobject_parents) == 0:
                    sheet_updates.append(
                        SheetRowUpdate(
                            records="", 
                            messages=",".join(messages)
                        )
                    )
                    continue
                for parent in sobject_parents:
                    type = parent['attributes']['type']
                    tree_url = f"composite/tree/{type}/"
                    print(tree_url)
                    await self.log(f"Inserting {type} object tree\n", callbacks)
                    pprint(parent)
                    data = json.dumps({"records": [parent]})
                    await self.log("Records:\n", callbacks)
                    await self.log(data+"\n", callbacks)
                    print(data)
                    try:
                        response = await sf.restful(
                            tree_url, 
                            method='POST',
                            data=data)
                    except SalesforceError as e:
                        response = e.content
                    print(response)
                    res: dict = response
                    if 'message' in res:
                        messages.append(res['message'])
                    records = []
                    errors = messages.copy()
                    if 'results' in res:
                        records = [result['id'] for result in res['results'] if 'id' in result]
                        records.extend([result['Id'] for result in res['results'] if 'Id' in result])
                        errors.extend([
                            result['errors'][0]['message'] for result in res['results'] if 'errors' in result
                        ])
                    inserted_count += len(records)
                    sheet_updates.append(
                        SheetRowUpdate(
                            records=",".join(records), 
                            messages=",".join(errors)
                        )
                    )
                    await self.log("Result was:" + str(sheet_updates[-1]), callbacks)
                    final_results.append(response)
                
        return f"Success: insert {inserted_count} records"

    @staticmethod 
    def update_sheet_results(
        excel_source_file: str, 
        sheet,
        workbook,
        sheet_updates: list[SheetRowUpdate]
        ):
        """
        Update the Excel sheet with the results of the insert operation.

        Parameters:
        excel_source_file (str): The path to the Excel file.
        sheet: The worksheet to update.
        workbook: The workbook containing the sheet.
        sheet_updates (list[SheetRowUpdate]): A list of updates to apply to the sheet.

        Returns:
        str: Success message after updating the Excel sheet.
        """
        next_col = sheet.max_column + 1
        sheet.cell(row=1, column=next_col, value="Uploaded Records")

        for row, data in enumerate(sheet_updates, start=2):  # skip the header
            sheet.cell(row=row, column=next_col, value=data.records)
        next_col += 1
        sheet.cell(row=1, column=next_col, value="Messages")
        for row, data in enumerate(sheet_updates, start=2):  # skip the header
            sheet.cell(row=row, column=next_col, value=data.messages)
        # Save the workbook
        workbook.save(excel_source_file)
        return "Success"

    @staticmethod
    def map_target_objects(headers: list, sheet, mappings: list, start_row: int, max_rows: int|None=None):
        """
        Map the target Salesforce objects based on the provided Excel sheet and mappings.

        Parameters:
        headers (list): A list of column headers from the Excel sheet.
        sheet: The worksheet object containing the data.
        mappings (list): A list of field mappings.
        start_row (int): The row number to start reading from.
        max_rows (int|None, optional): Maximum number of rows to process. Defaults to None.

        Yields:
        list[dict], int, list[str]: The mapped Salesforce objects, row index, and any messages.
        """
        headers = [h.lower() for h in headers]
        if max_rows:
            max_rows += 1 # offset for header
        for row_index, row in enumerate(sheet.iter_rows(min_row=(2+start_row), max_row=max_rows, values_only=True)):
            row_dict = {header.strip().replace('\n', '').replace('\r', ''): value for header, value in zip(headers, row)}

            targets = {}
            target_list: list[dict] = []
            refNumber = 1
            messages = []
            last_field:str = None

            for mapping in mappings:
                required = ((mapping.required or "").lower() == "true")
                target: dict
                target_object, target_field = mapping.target_field.split(".")
                target_field = target_field.strip()
                if target_object not in targets or (target_field in targets[target_object] and target_field != last_field):
                    target = {
                        "attributes": {
                            "type": target_object,
                            "referenceId": f"ref{refNumber}"
                        }
                    }
                    target_list.append(target)
                    refNumber += 1
                    targets[target_object] = target
                else:
                    target = targets[target_object]
                if mapping.source_col:
                    if mapping.source_col.lower() not in row_dict:
                        messages.append(f"Skipping object {target_object}, unknown column '{mapping.source_col}'")
                        target['__skip'] = True
                    else:
                        value = row_dict[mapping.source_col.lower()]
                    if mapping.target_val and "%%" in mapping.target_val:
                        value = mapping.target_val.replace("%%", value or "")
                    if not value and required:
                        messages.append(f"Skipping object {target_object} missing required field '{mapping.source_col}'")
                        target['__skip'] = True
                else:
                    value = mapping.target_val
                if isinstance(value, datetime):
                    value = value.isoformat() + 'Z'
                if target_field:
                    if target_field in target:
                        target[target_field] += "\n" + value
                    else:
                        target[target_field] = value
                    last_field = target_field
                if mapping.parent_object:
                    target['__parent'] = mapping.parent_object
                    parentId = mapping.parent_object + "Id"
                    if parentId in target:
                        del target[parentId]

            target_list = [t for t in target_list if '__skip' not in t]
            yield target_list, row_index, messages


    async def upsert_list_of_objects(
        self, 
        sobject: str,
        dataframe_var: str,
        lookup_column: str,
        max_records: int|None=None,
    ):
        """
        Upsert records from a dataframe into the specified Salesforce object.

        Parameters:
        sobject (str): The Salesforce object type.
        dataframe_var (str): The dataframe variable containing the records to upsert.
        lookup_column (str): The column to use for matching existing records.
        max_records (int|None, optional): Maximum number of records to process. Defaults to None.
        callbacks (LangChainCallback, optional): Optional callbacks for logging.

        Returns:
        str: A summary of the upsert operation.
        """
        await self.log(f"Performing upsert for {sobject} using {lookup_column}\n")
        batch_size = 200
        if max_records:
            batch_size = min(max_records, batch_size)

        target_fields = await self.get_object_fields(sobject)
        lookup_column = self.fix_column_name(lookup_column, target_fields)

        async with self.sf() as sf:
            all_records, _ = self.get_dataframe_from_handle(dataframe_var)

            renames = {}
            for col in all_records.columns:
                new_col = self.fix_column_name(col, target_fields)
                if new_col != col:
                    renames[col] = new_col
            if renames:
                all_records.rename(columns=renames, inplace=True)

            for df in self.dataframe_batch_iterator(all_records, batch_size):
                existing_records = {}
                key_list = df[lookup_column].tolist()
                query = f"SELECT Id, {lookup_column} FROM {sobject} WHERE {lookup_column} IN {tuple(key_list)}"
                await self.log(f"Query: {query}\n")
                results = await sf.query_all(query)
                await self.log(f"Results: {results}\n")
                for record in results['records']:
                    existing_records[record[lookup_column]] = record['Id']

                records_to_update = []
                records_to_insert = []

                for _, row in df.iterrows():
                    record_data = row.to_dict()
                    keyval = record_data[lookup_column]
                    
                    if keyval in existing_records:
                        record_data['Id'] = existing_records[keyval]
                        records_to_update.append(record_data)
                    else:
                        records_to_insert.append(record_data)

                messages = []
                if records_to_update:
                    await self.log("Updating records via Bulk API")
                    update_result = await getattr(sf.bulk, sobject).update(records_to_update)
                    if is_async_generator(update_result):
                        update_result = [r async for r in update_result]
                    messages.append(f"Updated {len(update_result)} existing Lead records")

                if records_to_insert:
                    await self.log("Inserting records via Bulk API")
                    insert_result = await getattr(sf.bulk, sobject).insert(records_to_insert)
                    messages.append(f"Inserted {len(insert_result)} new Lead records")

                def check_results(results, operation, errors: list):
                    for result in results:
                        if not result['success']:
                            errors.append(f"Error in {operation}: {result['errors']}")

                errors = []
                if records_to_update:
                    check_results(update_result, "update", errors)
                if records_to_insert:
                    check_results(insert_result, "insert", errors)

                return "\n".join(messages) + "\n" + "\n".join(errors)

    def fix_column_name(self, col: str, target_fields: list[str]) -> str:
        """
        Corrects a column name to match Salesforce field names, with case-insensitive matching and Levenshtein distance for close matches.

        Parameters:
        col (str): The column name to correct.
        target_fields (list[str]): List of Salesforce field names to compare against.

        Returns:
        str: The corrected column name, or the original name if no close match is found.
        """
        for testcol in target_fields:
            if testcol == col:
                return col
        for testcol in target_fields:      
            if testcol.lower() == col.lower():
                return testcol
        for testcol in target_fields:      
            if testcol.lower() == col.lower():
                return testcol
        for testcol in target_fields:
            if distance(col.lower(), testcol.lower()) < 3:
                return testcol
        return col


    def metadata_api_create_custom_object(
            self, 
            fullName: str, 
            fields: list[dict],
            overwrite: bool=False) -> str:
        """
        Defines a new custom object in Salesforce using the Metadata API.

        Args:
            fullName (str): The full name of the custom object to create.
            fields (list[dict]): A list of dictionaries defining fields.
            overwrite (bool, optional): If True, overwrites an existing object. Defaults to False.

        When specifying a MasterDetail field, you MUST supply the referenceTo attribute.
        Fields should specify their name in the 'label' property.
        Use "Restrict" as the delete constraint for custom lookup relationships.

        Returns:
            str: A string describing the result of the custom object creation attempt.
        """
        with self.sf_synch() as sf:
            mdapi = sf.mdapi
            mdtype = mdapi.CustomObject

            if overwrite:
                try:
                    delete_result =  mdtype.delete([fullName])
                    print(f"Delete result: {delete_result}")
                except Exception as e:
                    print(f"Error during delete: {e}")
                    print(traceback.format_exc())

            sharing_model = mdapi.SharingModel("Read")
            try:
                # Prepare fields asynchronously
                prepared_fields = []
                for field in fields:
                    print(f"Preparing field: {field['label']}, type: {field['type']}")

                    # Use the field type directly, without attempting to resolve from WSDL
                    field_type_value = field['type']
                    print(f"Debug: Field type value passed to CustomField: {field_type_value}")
                    if field_type_value == "MasterDetail":
                        sharing_model = None

                    custom_field = mdapi.CustomField(**field)
                    prepared_fields.append(custom_field)

                # Prepare nameField
                name_field = mdapi.CustomField(
                    label="Name",
                    type="Text"  # Pass "Text" as the field type for nameField
                )

                # Create the custom object
                custom_object = mdtype(
                    fullName=fullName,
                    label=fullName,
                    pluralLabel=fullName,
                    deploymentStatus=mdapi.DeploymentStatus("Deployed"),
                    nameField=name_field,
                    fields=prepared_fields
                )
                if sharing_model:
                    custom_object.sharingModel = sharing_model

                res = f"Custom object created: {repr(custom_object)}\n"

                # Create the custom object asynchronously
                create_result =  mdtype.create([custom_object])
                res += str(create_result)
            except Exception as e:
                res = f"Error creating custom object: {str(e)}"
                print(traceback.format_exc())

            print(res)
            return res

    async def delete_custom_object(self, object_name: str, challenge: str, answer: str) -> str:
        """ Deletes a custom object from Salesforce. Before you call this function you must
            prompt the user to answer a short confirmation challenge. Then pass the requested challenge
            value and the user's matching answer to this function. """
        if challenge != answer:
            return "Requst not confirmed"
        
        with self.sf_synch() as sf:
            mdapi = sf.mdapi
            mdtype = mdapi.CustomObject

            res = mdtype.delete([object_name])
            return str(res)

    async def query_log_id_list_from_apex(self):
        """ Returns the most recent set of Apex execution Runs. """

        async with self.sf() as sf:
            query = "SELECT Id, LogLength, LastModifiedDate FROM ApexLog ORDER BY LastModifiedDate DESC LIMIT 10"
            res = await sf.toolingexecute("query", params={"q": query})
            return res['records']

    async def get_latest_apex_logs(self, prior_offset: int=0) -> str:
        """ Returns the Apex logs for the most recent Apex run, or you can provide a log_id
            to get logs for a specific run. Or provide an offset to retrieve logs for 
            latest-Nth run. """

        runs = await self.query_log_id_list_from_apex()
        if prior_offset:
            log_id = runs[prior_offset]['Id']
        else:
            log_id = runs[0]['Id']

        async with self.sf() as sf:
            path = f"tooling/sobjects/ApexLog/{log_id}/Body"
            url = sf.base_url + path
            result = await sf._call_salesforce("GET", url, name=path)
            lines = [l for l in result.text.split("\n") if 'STATEMENT_EXECUTE' not in l and 'HEAP_ALLOCATE' not in l]
            return "\n".join(lines)
            
      
    async def list_apex_classes(self, name_prefix: str="") -> list[dict]:
        """ Returns the list of all Apex classes. Optionally provide a name prefix to only return matching classes. """

        async with self.sf() as sf:
            where = f"WHERE Name LIKE '{name_prefix}%'" if name_prefix else ""
            query = f"SELECT Id, Name, ApiVersion FROM ApexClass {where} ORDER BY Name"
            res = await sf.toolingexecute("query", params={"q": query})
            return res['records']

    async def get_apex_class(self, class_name: str|None=None, class_id:str|None=None) -> str:
        """ Returns the source code to an Apex class, indentified by name or Id. """
            
        if class_name:
            matches = await self.list_apex_classes(class_name)
            class_id = matches[0]['Id']

        if class_id is None:
            return "Error, must provide class_id or a name which matches an ApexClass"
        
        return await self.get_object_by_id(class_id, "ApexClass")
    
      
        